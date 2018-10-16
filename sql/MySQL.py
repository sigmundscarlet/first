# ！/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
© Copyright 2018 The Author. All Rights Reversed.
------------------------------------------------------------
File Name: 
Author : zhangtao 
Time:
Description：
------------------------------------------------------------
"""
import re, os, sys
import pymysql.cursors
import openpyxl

import regex_openlaw as openlaw
import party_name
# import explanation_of_trail
def write_to_MySQL(path_openlaw, log):
    # connection = pymysql.connect(host='10.20.96.152',
    #                              user='myuser',
    #                              password='1234qwer',
    #                              db='openlawtest',
    #                              charset='utf8',
    #                              cursorclass=pymysql.cursors.DictCursor)

    # connection = pymysql.connect(host='localhost',
    #                              user='root',
    #                              password='yfzwfz2017-2020',
    #                              db='openlaw_test',
    #                              charset='utf8',
    #                              cursorclass=pymysql.cursors.DictCursor)

    #
    connection = pymysql.connect(host='10.20.96.152',
                                 user='myuser',
                                 password='yfzwfz2017-2020',
                                 db='openlaw',
                                 charset='utf8',
                                 cursorclass=pymysql.cursors.DictCursor)


    try:
        wb = openpyxl.load_workbook(path_openlaw)  # 打开excel文件

        sheet=wb.worksheets[0]  #获取工作表


        log.write('读取判决文书数据: '+ str(sheet.max_row - 1) + "条。\n\n")

        update_number = 0

        for r in range(2, sheet.max_row + 1):
            title = sheet.cell(row=r, column=1).value
            case_number = sheet.cell(row=r, column=2).value

            with connection.cursor() as cursor:
                sql = "SELECT OpenLawData.`案号` FROM OpenLawData WHERE `案号` = '%s'" % (case_number)
                cursor.execute(sql)
                results = cursor.fetchall()
            connection.commit()
            if (results):
                continue
            else:
                update_number += 1

                case_type = sheet.cell(row=r, column=3).value
                procedure_of_trial = sheet.cell(row=r, column=4).value

                case_of_action = openlaw.split_text(sheet.cell(row=r, column=5).value)

                document_type = sheet.cell(row=r, column=6).value
                court = sheet.cell(row=r, column=7).value
                judgment_date = sheet.cell(row=r, column=8).value
                judgment_date = re.findall('\d{1,4}', judgment_date)

                if len(judgment_date) == 3:
                    judgment_date = judgment_date[0] + '-' + judgment_date[1] + '-' + judgment_date[1]
                else:
                    judgment_date = None

                plaintiff = sheet.cell(row=r, column=9).value
                plaintiff = openlaw.split_text(plaintiff)

                defendant = sheet.cell(row=r, column=10).value
                defendant = openlaw.split_text(defendant)

                judge = sheet.cell(row=r, column=12).value
                presiding_judge = sheet.cell(row=r, column=13).value

                judicial_officer = sheet.cell(row=r, column=14).value
                judicial_officer = openlaw.split_text(judicial_officer)

                court_clerk = sheet.cell(row=r, column=15).value

                party = sheet.cell(row=r, column=18).value
                explanation_trail = sheet.cell(row=r, column=22).value

                # insert openlaw_ie table

                with connection.cursor() as cursor:
                    sql = """INSERT IGNORE INTO `openlaw_ie` (`case_number`, `title`, `case_type`, `procedure_of_court_trial`,`document_type`, `court`, `judgment_date`, `judge`, `presiding_judge`, `court_clerk`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
                    cursor.execute(sql, (case_number, title, case_type, procedure_of_trial, document_type, court, judgment_date, judge, presiding_judge, court_clerk))
                connection.commit()

                # insert case_of_action table
                if case_of_action == ['']:
                    pass
                else:
                    for i in range(len(case_of_action)):

                        with connection.cursor() as cursor:
                            # Create a new record
                            sql = "INSERT IGNORE INTO `case_of_action` (`case_number`, `case_of_action`) VALUES (%s, %s)"
                            cursor.execute(sql, (case_number, case_of_action[i]))
                        connection.commit()

                # insert defendant table
                if defendant == ['']:
                    pass
                else:
                    for i in range(len(defendant)):

                        with connection.cursor() as cursor:
                            # Create a new record
                            sql = "INSERT IGNORE INTO `defendant` (`case_number`, `defendant`) VALUES (%s, %s)"
                            cursor.execute(sql, (case_number, defendant[i]))
                        connection.commit()


                # insert explanation_of_trail table

                # insert judicial_officer table
                if judicial_officer == ['']:
                    pass
                else:
                    for i in range(len(judicial_officer)):

                        with connection.cursor() as cursor:
                            sql = "INSERT IGNORE INTO `judicial_officer` (`case_number`, `judicial_officer`) VALUES (%s, %s)"
                            cursor.execute(sql, (case_number, judicial_officer[i]))
                        connection.commit()
                # insert party_name table

                # insert plaintiff table
                if plaintiff == ['']:
                    pass
                else:
                    for i in range(len(plaintiff)):

                        with connection.cursor() as cursor:
                            sql = "INSERT IGNORE INTO `plaintiff` (`case_number`, `plaintiff`) VALUES (%s, %s)"
                            cursor.execute(sql, (case_number, plaintiff[i]))
                        connection.commit()

                list_temp = party_name.party_IE(party)
                if list_temp == [None, None, None, None, None, None, None, None, None]:
                    pass
                else:
                    for list_temp_ in list_temp:
                        name, people, province, gender, education, birth_palce, birthday, laywer, law_firm = list_temp_


                        with connection.cursor() as cursor:
                            sql = """INSERT IGNORE INTO `party_name` (`case_number`, `name`, `people`, `province`,\
                            `gender`, `education`, `birth place`, `birthday`, `laywer`, `law_firm`) VALUES \
                            (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
                            cursor.execute(sql, (case_number, name, people, province, gender, education, birth_palce, \
                                                 birthday, laywer, law_firm))
                        connection.commit()
        log.write('本次更新数据: ' + str(update_number) + "条。\n\n")

    finally:
        connection.close()