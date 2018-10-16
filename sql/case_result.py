# ！/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
© Copyright 2018 The Author. All Rights Reversed.
------------------------------------------------------------
File Name: 
Author : zhangtao 
Time:
Description：
------------------------------------------------------------
"""

import openpyxl
import re

import regex_openlaw
def result(path_openlaw):

    wb = openpyxl.load_workbook(path_openlaw)  # 打开excel文件

    sheet = wb.worksheets[0]  # 获取工作表
    print(sheet.title)

    for r in range(2, 100):
        text = sheet.cell(row=r, column=30).value
        text_Court_opinion = sheet.cell(row=r, column=28).value

        print('判决结果', regex_openlaw.get_judgment_result(text))
        string = regex_openlaw.get_legal_basis(text)
        if len(string) != 0:
            print('法律依据', re.split('、', string[0]))
        print('是否自首',regex_openlaw.get_is_surrendere(text_Court_opinion))

        print('法院意见', regex_openlaw.get_court_opinion(text_Court_opinion))
